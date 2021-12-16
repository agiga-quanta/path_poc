from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

class ExcelExporter(object):

    def __init__(self, config):
        pass

    def export(self, doc_dict, lookup, reason, file_name):
        workbook = Workbook()

        sheet = workbook.active
        sheet.title = 'Extractions'

        centered_aligment = Alignment(vertical='center')
        right_centered_aligment = Alignment(horizontal='right', vertical='center')
        middle_centered_aligment = Alignment(horizontal='center', vertical='center')
        wrapped_centered_aligment = Alignment(vertical='center', wrap_text=True)

        sheet['A1'] = 'File No'
        sheet['A1'].font = Font(bold=True)
        sheet['A1'].alignment = middle_centered_aligment
        sheet['B1'] = 'Proponent'
        sheet['B1'].font = Font(bold=True)
        sheet['B1'].alignment = middle_centered_aligment
        sheet['C1'] = 'Location'
        sheet['C1'].font = Font(bold=True)
        sheet['C1'].alignment = middle_centered_aligment
        sheet['D1'] = 'Impact Type'
        sheet['D1'].font = Font(bold=True)
        sheet['D1'].alignment = middle_centered_aligment
        sheet['E1'] = 'Footprint'
        sheet['E1'].font = Font(bold=True)
        sheet['E1'].alignment = middle_centered_aligment
        sheet['F1'] = 'Units'
        sheet['F1'].font = Font(bold=True)
        sheet['F1'].alignment = middle_centered_aligment
        sheet['G1'] = '(cont.)'
        sheet['G1'].font = Font(bold=True)
        sheet['G1'].alignment = middle_centered_aligment
        sheet['H1'] = 'Key Impact'
        sheet['H1'].font = Font(bold=True)
        sheet['H1'].alignment = middle_centered_aligment
        sheet['I1'] = 'Impact Sentence'
        sheet['I1'].font = Font(bold=True)
        sheet['I1'].alignment = middle_centered_aligment

        line_no = 2
        for path_id, doc_info in doc_dict.items():
            proponent, location, desc_list = doc_info
            doc_line_no = line_no
            sheet['A%s' % line_no] = path_id
            sheet['A%s' % line_no].alignment = centered_aligment
            sheet['B%s' % line_no] = '\n'.join('%s: %s' %(k.upper(), ', '.join(v)) for k, v in proponent.items() if k != 'nlp')
            sheet['B%s' % line_no].alignment = centered_aligment
            sheet['C%s' % line_no] = '\n'.join('%s: %s' %(k.upper(), ', '.join([e['t'] for e in v])) for k, v in location.items())
            sheet['C%s' % line_no].alignment = centered_aligment
            for sent_dict in desc_list:
                sheet['I%s' % line_no] = sent_dict['sent']
                sheet['I%s' % line_no].alignment = wrapped_centered_aligment
                if 'impact' in sent_dict:
                    sheet['D%s' % line_no] = ';\n'.join(', '.join(e) for e in sent_dict['impact'].values())
                    sheet['D%s' % line_no].alignment = centered_aligment
                if 'reason' in sent_dict:
                    sheet['H%s' % line_no] = ';\n'.join(list(set(', '.join(e) for e in sent_dict['reason'].values())))
                    sheet['H%s' % line_no].alignment = centered_aligment
                sent_line_no = line_no

                if 'footprint' in sent_dict and sent_dict['footprint']:
                    for _, wf_wr in sent_dict['footprint'].items():
                        wf, wr = wf_wr
                        sheet['E%s' % line_no] = wf
                        sheet['E%s' % line_no].alignment = right_centered_aligment
                        sheet['F%s' % line_no] = wr[0]
                        sheet['F%s' % line_no].alignment = middle_centered_aligment
                        sheet['G%s' % line_no] = ' '.join(wr[1:])
                        sheet['G%s' % line_no].alignment = centered_aligment
                        line_no += 1
                else:
                    line_no += 1

                if sent_line_no <= line_no-1:
                    for letter in ['D', 'I', 'H']:
                        sheet.merge_cells('%s%s:%s%s' % (letter, sent_line_no, letter, line_no-1))
            
            if doc_line_no <= line_no-1:
                for letter in ['A', 'B', 'C']:
                    sheet.merge_cells('%s%s:%s%s' % (letter, doc_line_no, letter, line_no-1))
            line_no += 1

        sheet = workbook.create_sheet(title='Keywords')
        sheet['A1'] = 'Keywords'
        sheet['B1'] = lookup
        sheet['A2'] = 'Reasons'
        sheet['B2'] = reason

        workbook.save(filename=file_name)